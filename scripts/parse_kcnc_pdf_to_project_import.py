from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
PDF_PATH = Path(r"C:\Users\nguye\Downloads\2026-01-05_-_DT_QDUB_DM_DT_KCNC_signed (1)(1).pdf")
TEMPLATE_PATH = ROOT / "docs/ba/operation/templates/HCMInvHub_Project_Import_Template_v1_vi_accented2.xlsx"
OUTPUT_DIR = ROOT / "outputs/project_import_kcnc_20260105"
OUTPUT_PATH = OUTPUT_DIR / "project_import_kcnc_20260105.xlsx"
EXTRACT_TEXT_PATH = OUTPUT_DIR / "pdf_extract_kcnc_20260105.txt"
SHTP_AGENCY_PDF_VALUE = "Ban Quản lý Khu Công nghệ cao Thành phố Hồ Chí Minh"


def clean_text(value: str) -> str:
    value = re.sub(r"===== PAGE \d+ =====", " ", value)
    value = re.sub(r"\n\s*\d+\s*\n", "\n", value)
    value = value.replace("độ", "độ").replace("mộ", "mộ").replace("lộ", "lộ")
    value = value.replace("tiê", "tiê").replace("Tên", "Tên")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{2,}", "\n", value)
    return value.strip()


def one_line(value: str | None, limit: int | None = None) -> str:
    if not value:
        return ""
    value = clean_text(value)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:limit].rstrip() if limit else value


def extract_between(block: str, start_pattern: str, end_patterns: list[str], limit: int | None = None) -> str:
    start = re.search(start_pattern, block, flags=re.I | re.S)
    if not start:
        return ""
    sub = block[start.end() :]
    end_positions = []
    for pattern in end_patterns:
        end = re.search(pattern, sub, flags=re.I | re.S)
        if end:
            end_positions.append(end.start())
    if end_positions:
        sub = sub[: min(end_positions)]
    return one_line(sub, limit)


def parse_vn_number(value: str) -> float | None:
    match = re.search(r"\d[\d.,]*", value)
    if not match:
        return None
    raw = match.group(0)
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "." in raw and re.search(r"\.\d{3}(?:$|\D)", raw):
        raw = raw.replace(".", "")
    try:
        return float(raw)
    except ValueError:
        return None


def extract_area_m2(block: str) -> tuple[float | None, str]:
    match = re.search(r"Diện tích:\s*([^\n]+)", block, flags=re.I)
    if not match:
        return None, ""
    raw = match.group(1).strip()
    return parse_vn_number(raw), one_line(raw)


def extract_usd_per_ha(block: str) -> float | None:
    values = []
    for match in re.finditer(r"(\d[\d.,]*)\s*đô la\s*M\s*ỹ\s*/\s*ha", block, flags=re.I):
        parsed = parse_vn_number(match.group(1))
        if parsed:
            values.append(parsed)
    return values[-1] if values else None


def extract_investment_rate_text(block: str) -> str:
    match = re.search(r"Vốn[^\n]*([\s\S]{0,450}?)(?=(?:Chỉ tiêu|Thông tin|Các chỉ tiêu|$))", block, flags=re.I)
    if not match:
        return ""
    return one_line(match.group(0))


def extract_address(block: str, lot: str) -> str:
    patterns = [
        rf"-\s*(?:Địa điểm:\s*)?({re.escape(lot)}[\s\S]{{0,260}}?Thành phố Hồ Chí Minh(?:, Việt Nam)?)",
        r"-\s*(?:Địa điểm:\s*)?(Lô[\s\S]{0,260}?Thành phố Hồ Chí Minh(?:, Việt Nam)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, block, flags=re.I)
        if match:
            return one_line(match.group(1))
    return f"{lot}, Khu Công nghệ cao, phường Tăng Nhơn Phú, Thành phố Hồ Chí Minh"


def classify_sector(text: str) -> tuple[str, str]:
    lower = text.lower()
    if any(word in lower for word in ["dữ liệu", "điện toán", "vi mạch", "bán dẫn", "điện tử"]):
        main = "High-Tech & Digital"
    elif any(word in lower for word in ["sinh học", "y sinh", "vắc xin", "liệu pháp"]):
        main = "Biotech & Life Sciences"
    else:
        main = "High-Tech & Digital"

    subs = []
    if any(word in lower for word in ["sản xuất", "nhà máy"]):
        subs.append("Manufacturing")
    if any(word in lower for word in ["sinh học", "y sinh", "vắc xin", "liệu pháp"]):
        subs.append("Biotech & Life Sciences")
    if any(word in lower for word in ["điện tử", "vi mạch", "bán dẫn", "dữ liệu", "điện toán"]):
        subs.append("High-Tech & Digital")
    if "đào tạo" in lower:
        subs.append("Education")
    return main, "; ".join(dict.fromkeys(subs))


def map_pdf_sector_to_database_sector(lot: str, sector_text: str, project_type_text: str) -> tuple[str, str, str]:
    source = f"{sector_text} {project_type_text}".lower()
    secondary: list[str] = []

    has_semiconductor = any(token in source for token in ["vi mạch", "bán dẫn", "điện tử", "dữ liệu", "điện toán", "trí tuệ nhân tạo", "data center"])
    has_biotech = any(token in source for token in ["sinh học", "y sinh", "vắc xin", "liệu pháp", "dược"])
    has_manufacturing = any(token in source for token in ["nhà máy", "sản xuất"])
    has_training = "đào tạo" in source

    if has_semiconductor:
        primary = "High-Tech & Digital"
    elif has_biotech:
        primary = "Biotech & Life Sciences"
    else:
        primary = ""

    if has_semiconductor:
        secondary.append("High-Tech & Digital")
    if has_biotech:
        secondary.append("Biotech & Life Sciences")
    if has_manufacturing:
        secondary.append("Manufacturing")
    if has_training:
        secondary.append("Education")

    secondary = [item for item in dict.fromkeys(secondary) if item != primary]
    note = f"Mapped from PDF priority sector text: {sector_text}" if primary else ""
    return primary, "; ".join(secondary), note


def project_name(lot: str, sector: str) -> str:
    if lot == "Lô T7":
        return "Trung tâm dịch vụ khoa học - công nghệ cao đa chức năng tại Lô T7"
    if "dữ liệu" in sector.lower():
        return f"Trung tâm dữ liệu tại {lot}"
    if "dịch vụ khoa học" in sector.lower() or "dịch vụ công nghệ cao" in sector.lower():
        return f"Trung tâm dịch vụ khoa học - công nghệ cao tại {lot}"
    if "sản xuất" in sector.lower() or "nhà máy" in sector.lower():
        return f"Nhà máy sản xuất công nghệ cao tại {lot}"
    return f"Dự án nghiên cứu phát triển - đào tạo tại {lot}"


def extract_project_type(block: str) -> str:
    lower = block.lower()
    if "đất dành cho hoạt độ" in lower and "nghiên cứu phát triển - đào tạo" in lower:
        return "Nghiên cứu phát triển - Đào tạo"
    if "đất dành cho dự án sản xuất công nghệ cao" in lower:
        return "Sản xuất công nghệ cao"
    if "đất sử dụng vào hoạt độ" in lower and "dịch vụ công nghệ cao" in lower:
        return "Dịch vụ công nghệ cao"
    return ""


def extract_contact(text: str) -> dict[str, str]:
    section = extract_between(text, r"6\.\s*Đầu mối liên hệ", [r"B\."], 1000)
    email = re.search(r"Email:\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})", section)
    phone = re.search(r"Điện thoại:\s*([^\n.]+)", section)
    website = re.search(r"Website:\s*([^\s.]+)", section)
    return {
        "name": "BAN QUẢN LÝ KHU CÔNG NGHỆ CAO THÀNH PHỐ HỒ CHÍ MINH" if "BAN QUẢN LÝ KHU CÔNG NGHỆ CAO" in section else "",
        "email": email.group(1).rstrip(".") if email else "",
        "phone": one_line(phone.group(1)) if phone else "",
        "website": website.group(1).rstrip(".") if website else "",
    }


def clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def remove_unproven_output_data(wb) -> None:
    main = wb[wb.sheetnames[2]]
    custom = wb[wb.sheetnames[5]]

    allowed_main_fields = {
        "mã_dự_án",
        "tên_cơ_quan_sở_hữu",
        "tên_dự_án",
        "tỉnh_thành",
        "phường_xã",
        "địa_chỉ_chi_tiết",
        "mô_tả",
        "người_liên_hệ",
        "email_liên_hệ",
        "số_điện_thoại_liên_hệ",
    }
    for col, header in enumerate([cell.value for cell in main[1]], start=1):
        if header not in allowed_main_fields:
            for row in range(2, main.max_row + 1):
                main.cell(row, col, value=None)

    allowed_custom_labels = {
        "Ký hiệu lô",
        "Diện tích gốc",
        "Mục đích sử dụng đất",
        "Loại hoạt động theo PDF",
        "Lĩnh vực ưu tiên thu hút theo PDF",
        "Thông tin vốn đầu tư theo PDF",
    }
    label_col = 3
    for row in range(custom.max_row, 1, -1):
        if custom.cell(row, label_col).value not in allowed_custom_labels:
            custom.delete_rows(row)

    for sheet_index in [3, 4, 6, 7, 8, 9]:
        clear_data_rows(wb[wb.sheetnames[sheet_index]])


def main() -> None:
    reader = PdfReader(str(PDF_PATH))
    text = "\n".join(f"\n===== PAGE {idx + 1} =====\n{page.extract_text() or ''}" for idx, page in enumerate(reader.pages))
    text = clean_text(text)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    EXTRACT_TEXT_PATH.write_text(text, encoding="utf-8")
    start = text.find("1.1")
    project_text = text[start:]
    heads = list(re.finditer(r"(?m)^\s*(\d+[.]\d+)\s+([^\n]+)", project_text))

    contact = extract_contact(text)
    legal_summary = extract_between(text, r"Căn cứ Quyết định số 4819/QĐ-UBND", [r"Điều 1"], 900)
    incentives_text = extract_between(text, r"3\.\s*Ưu đãi đầu tư", [r"4\.\s*Giá thuê đất"], 2000)
    investment_call = extract_between(text, r"2\.\s*Hình thức kêu gọi đầu tư", [r"3\.\s*Ưu đãi đầu tư"], 1200)

    projects = []
    for index, head in enumerate(heads):
        end = heads[index + 1].start() if index + 1 < len(heads) else len(project_text)
        block = project_text[head.start() : end]
        block = clean_text(block)
        lot = one_line(head.group(2))
        area_m2, area_text = extract_area_m2(block)
        area_ha = round(area_m2 / 10000, 4) if area_m2 else None
        sector_text = extract_between(block, r"Lĩnh vực ưu[^\n]*\n(?:thu hút\s*)?", [r"\nCông nghệ", r"\n- Dự án"], 650)
        if not sector_text:
            sector_text = extract_between(block, r"Lĩnh vực ưu[\s\S]{0,80}?thu hút", [r"\nCông nghệ", r"\n- Dự án"], 650)
        address = extract_address(block, lot)
        land_match = re.search(r"-\s*Đất[^\n]+", block)
        land_use = one_line(land_match.group(0).lstrip("- ")) if land_match else ""
        tech = extract_between(block, r"\nCông nghệ", [r"\nSản ph", r"\nVốn"], 2400)
        products = extract_between(block, r"Sản ph[\s\S]{0,80}nghệ cao", [r"\nVốn", r"\nChỉ tiêu", r"\nThông tin"], 1600)
        planning = extract_between(block, r"(?:Chỉ tiêu quy|Thông tin quy)[\s\S]{0,80}?khu đất", [r"\n\d+[.]\d+\s+", r"\nC\.\s", r"\nD\.\s"], 2600)
        if not planning:
            planning = extract_between(block, r"(?:Chỉ tiêu quy|Thông tin quy)", [r"\n\d+[.]\d+\s+", r"\nC\.\s"], 2600)

        project_type_text = extract_project_type(block)
        mapped_sector, mapped_secondary, sector_note = map_pdf_sector_to_database_sector(lot, sector_text, project_type_text)

        projects.append(
            {
                "id": head.group(1),
                "lot": lot,
                "name": "Tên dự án do nhà đầu tư đề xuất",
                "address": address,
                "area_text": area_text,
                "area_ha": area_ha,
                "land_use": land_use,
                "sector_text": sector_text,
                "project_type_text": project_type_text,
                "mapped_sector": "",
                "mapped_secondary": "",
                "sector_note": sector_note,
                "investment_rate_text": extract_investment_rate_text(block),
                "tech": tech,
                "products": products,
                "planning": planning,
            }
        )

    for project in projects:
        if project["lot"] == "Lô I-14.4":
            project["address"] = "Lô I-14.4, Đường D14, Khu Công nghệ cao, phường Tăng Nhơn Phú, Thành phố Hồ Chí Minh"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    ws_main = wb[wb.sheetnames[2]]
    ws_map = wb[wb.sheetnames[3]]
    ws_tasks = wb[wb.sheetnames[4]]
    ws_custom = wb[wb.sheetnames[5]]
    ws_docs = wb[wb.sheetnames[7]]
    ws_incentives = wb[wb.sheetnames[8]]

    for ws in [ws_main, ws_map, ws_tasks, ws_custom, ws_docs, ws_incentives]:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    decision_no = "Quyết định sửa đổi, cập nhật Quyết định số 4819/QĐ-UBND ngày 29/10/2024"
    decision_date = None
    contact_name = contact["name"]
    contact_email = contact["email"]
    contact_phone = contact["phone"]
    owner = SHTP_AGENCY_PDF_VALUE if contact["name"] else ""

    for row_idx, p in enumerate(projects, 2):
        description = p["sector_text"] or p["land_use"] or summary
        ws_main.append(
            [
                p["id"],
                owner,
                "Tên dự án do nhà đầu tư đề xuất",
                p["mapped_sector"],
                p["mapped_secondary"],
                "",
                "Thành phố Hồ Chí Minh",
                "",
                "Phường Tăng Nhơn Phú",
                p["address"],
                description,
                "",
                None,
                None,
                "",
                "",
                "",
                decision_no,
                decision_date,
                "",
                "",
                p["area_ha"],
                None,
                "",
                "",
                None,
                None,
                contact_name,
                contact_email,
                contact_phone,
            ]
        )
        ws_map.append([p["id"], None, None, p["lot"], p["address"], None])
        ws_docs.append([p["id"], "Quyết định danh mục dự án thu hút đầu tư Khu Công nghệ cao 2025-2030", "Legal", str(PDF_PATH), "Nguồn PDF do người dùng cung cấp; thông tin được trích xuất để nhập template.", 1])
        if incentives_text:
            ws_incentives.append([p["id"], "", incentives_text, "", 1])
        custom_rows = [
            ("Thông tin lô đất", "Ký hiệu lô", p["lot"]),
            ("Thông tin lô đất", "Diện tích gốc", p["area_text"]),
            ("Thông tin lô đất", "Mục đích sử dụng đất", p["land_use"]),
            ("Thông tin dự án", "Loại hoạt động theo PDF", p["project_type_text"]),
            ("Thông tin dự án", "Lĩnh vực ưu tiên thu hút theo PDF", p["sector_text"]),
            ("Ánh xạ Sector", "Ghi chú ánh xạ lĩnh vực", p["sector_note"]),
            ("Vốn đầu tư", "Thông tin vốn đầu tư theo PDF", p["investment_rate_text"]),
            ("Thông tin chung", "Hình thức kêu gọi đầu tư", investment_call),
            ("Pháp lý", "Căn cứ quyết định", legal_summary),
            ("Công nghệ ưu tiên", "Công nghệ", p["tech"]),
            ("Sản phẩm công nghệ cao", "Sản phẩm ưu tiên", p["products"]),
            ("Quy hoạch và hiện trạng", "Chỉ tiêu quy hoạch / hiện trạng", p["planning"]),
        ]
        order = 1
        for group, label, value in custom_rows:
            if value:
                ws_custom.append([p["id"], group, label, value, order])
                order += 1

    for ws in [ws_main, ws_map, ws_tasks, ws_custom, ws_docs, ws_incentives]:
        ws.freeze_panes = "A2"

    remove_unproven_output_data(wb)
    wb.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")
    print(f"Projects: {len(projects)}")
    for p in projects:
        print(p["id"], p["lot"], p["name"], p["area_ha"])


if __name__ == "__main__":
    main()
